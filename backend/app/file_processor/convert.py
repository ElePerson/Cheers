"""文件转 Markdown：txt/md/docx（M1）；pdf/xlsx（M2）；图片为占位。"""
from pathlib import Path


def _pdf_to_markdown(path: Path) -> str:
    """PDF 文本提取为 Markdown（PyMuPDF）。"""
    import pymupdf
    doc = pymupdf.open(path)
    parts = []
    for page in doc:
        parts.append(page.get_text())
    doc.close()
    text = "\f".join(parts).replace("\f", "\n\n")
    return text.strip() or "(无文本内容)"


def _xlsx_to_markdown(path: Path) -> str:
    """xlsx 转 Markdown 表格（openpyxl）。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"## {sheet.title}\n")
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        # 表头
        header = "| " + " | ".join(str(c) if c is not None else "" for c in rows[0]) + " |"
        sep = "| " + " | ".join("---" for _ in rows[0]) + " |"
        parts.append(header)
        parts.append(sep)
        for row in rows[1:21]:
            parts.append("| " + " | ".join(str(c) if c is not None else "" for c in row) + " |")
        if len(rows) > 21:
            parts.append(f"\n*（仅前 21 行，共 {len(rows)} 行）*")
        parts.append("")
    wb.close()
    return "\n".join(parts).strip() or "(空表)"


def _image_placeholder(_path: Path) -> str:
    """图片暂不调用 Vision API，返回占位说明。"""
    return "[图片内容描述需配置系统 Vision API 后生成]"


def to_markdown(file_path: str | Path) -> str:
    """
    将支持的文件转为 Markdown 文本。
    支持：.txt, .md 直接读取；.docx 用 mammoth；.pdf 用 pymupdf；.xlsx 用 openpyxl；
    .png/.jpg/.jpeg/.webp 返回占位说明。
    不支持则抛出 ValueError。
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(str(path))
    suffix = path.suffix.lower()
    if suffix in (".txt", ".md"):
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix == ".docx":
        import mammoth
        with open(path, "rb") as f:
            result = mammoth.convert_to_markdown(f)
        return result.value
    if suffix == ".pdf":
        return _pdf_to_markdown(path)
    if suffix == ".xlsx":
        return _xlsx_to_markdown(path)
    if suffix in (".png", ".jpg", ".jpeg", ".webp"):
        return _image_placeholder(path)
    raise ValueError(f"unsupported format: {suffix}")
